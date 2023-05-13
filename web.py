import time
import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

ultima_solicitud = time.time()
rate_interval = 2

headers = {
    'User-Agent': 'Usuario: DannaMedina, Escuela: UniversidadDeGuadalajara, Correo: danna.medina2869@alumnos.udg.mx'
}


def solicitar_tema():
    tema = input("Ingrese el tema a buscar en GitHub: ")
    return tema

def solicitar_lenguaje():
    lenguajes = {
        1: "Java",
        2: "Python",
        3: "JavaScript",
        4: "C++",
        5: "Ruby",
        6: "PHP",
        7: "C#",
        8: "Objective-C",
        9: "Swift",
        10: "Kotlin"
    }

    print("Seleccione el lenguaje de programación:")
    for num, lenguaje in lenguajes.items():
        print(f"{num}. {lenguaje}")

    num_lenguaje = int(input("Ingrese el número del lenguaje de programación: "))
    lenguaje_seleccionado = lenguajes.get(num_lenguaje)
    return lenguaje_seleccionado


def buscar_repositorios(tema, lenguaje):
    global ultima_solicitud
    hora_actual = time.time()
    tiempo_adv = hora_actual - ultima_solicitud
    if tiempo_adv < rate_interval:
        time.sleep(rate_interval - tiempo_adv)
    ultima_solicitud = time.time()
    
    url = f"https://api.github.com/search/repositories?q=topic:{tema}+language:{lenguaje}&sort=stars&order=desc"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        results = []
        data = response.json()
        for item in data['items'][:5]:
            results.append((item['owner']['login'], item['name'], item['html_url'], item['stargazers_count']))
        return results
    else:
        raise Exception("Error al hacer la solicitud: " + str(response.status_code))


def guardar_resultados(repositorios):
    wb = Workbook()
    ws = wb.active
    font = Font(name="Century Gothic")

    encabezados = ["Usuario", "Repositorio", "URL", "Puntuación"]
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.font = font
        ws.column_dimensions[chr(64 + col)].width = len(encabezado) + 2

    fila = 2
    for repositorio in repositorios:
        usuario, nombre, url, puntuacion = repositorio
        ws.cell(row=fila, column=1, value=usuario)
        ws.cell(row=fila, column=2, value=nombre)
        ws.cell(row=fila, column=3, value=url)
        ws.cell(row=fila, column=4, value=puntuacion)

        for col in range(1, 5):
            ws.cell(row=fila, column=col).font = font
            ws.column_dimensions[chr(64 + col)].width = max(ws.column_dimensions[chr(64 + col)].width,
                                                            len(str(ws.cell(row=fila, column=col).value)) + 2)

        fila += 1

    wb.save("Resultados.xlsx")

def verificar_resultados(repositorios):
    if len(repositorios) == 0:
        print("No se encontraron repositorios que coincidan con los criterios de búsqueda.")
    else:
        print(r"Los resultados fueron encontrados. El archivo 'Resultados.xlsx' ha sido guardado en la Ruta 'C:\Users\danna\Desktop\Resultados.xlsx'")
        abrir_archivo()

def abrir_archivo():
    # Abrir archivo
    wb = load_workbook("Resultados.xlsx")
    ws = wb.active

    # Guardar archivo antes de abrirlo
    wb.save("Resultados.xlsx")

    # Obtener la ruta absoluta del archivo
    ruta_absoluta = os.path.abspath("Resultados.xlsx")

    # Abrir el archivo en la aplicación predeterminada según el sistema operativo
    if os.name == "nt":  # Windows
        os.startfile(ruta_absoluta)
    elif os.name == "posix":  # Mac o Linux
        os.system("open " + ruta_absoluta)

    print("El archivo se ha abierto en tu aplicación predeterminada.")
    print("Recuerda cerrarlo manualmente cuando hayas terminado.")

def main():
    tema = solicitar_tema()
    lenguaje = solicitar_lenguaje()
    repositorios = buscar_repositorios(tema, lenguaje)
    guardar_resultados(repositorios)

    verificar_resultados(repositorios)

# Llamada a la función principal
main()
